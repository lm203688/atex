#!/usr/bin/env python3
"""ATEX综合日报生成器 - 2026-05-13"""

import sys
import os
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
from templates.base import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Initialize palette
use_palette_explicit("professional")
wb = Workbook()

# Custom fonts for title
TITLE_FONT = Font(name=CJK_BODY_CHAIN[0], size=16, bold=True, color="FFFFFF")
SECTION_FONT = Font(name=CJK_BODY_CHAIN[0], size=12, bold=True, color="2E75B6")
SUB_FONT = Font(name=CJK_BODY_CHAIN[0], size=9, color="666666")

DARK_FILL = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
ALT_FILL_1 = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
ALT_FILL_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(vertical="center", wrap_text=True)

def write_section_title(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = SECTION_FONT

def write_table(ws, start_row, headers, data, col_start=1):
    """Write a table with headers and data rows."""
    # Headers
    for c, h in enumerate(headers, col_start):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = font_header()
        cell.fill = fill_header()
        cell.alignment = align_header()
    # Data
    for idx, row_data in enumerate(data):
        r = start_row + 1 + idx
        for c, v in enumerate(row_data, col_start):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = font_body()
            cell.alignment = LEFT_WRAP
            cell.fill = fill_data_row(idx)
    return start_row + 1 + len(data)

# ============================================================
# Sheet1: 全球AI技术动态
# ============================================================
ws1 = wb.active
ws1.title = "全球AI技术动态"

# Title
ws1.merge_cells('A1:F1')
t1 = ws1['A1']
t1.value = "全球AI技术动态日报 — 2026年5月13日"
t1.font = TITLE_FONT
t1.fill = DARK_FILL
t1.alignment = CENTER_WRAP
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:F2')
s1 = ws1['A2']
s1.value = "数据来源：14组全球搜索 + 深度阅读 | 覆盖：OpenAI/Google/Anthropic/Meta/NVIDIA/中国大厂/融资/Agent协议/视频生成"
s1.font = SUB_FONT
s1.alignment = Alignment(horizontal="center")

# News data
news_data = [
    [1, "GPT-5.5 Instant发布", "OpenAI发布新一代基础模型GPT-5.5 Instant，替代GPT-5.3成为ChatGPT默认模型。准确性大幅提升，幻觉显著减少，法律和金融领域表现突出，延迟更低", "OpenAI", "2026-05-05", "⭐⭐⭐⭐⭐ 重新定义AI助手基准，法律/金融等专业领域可用性跃升"],
    [2, "OpenAI完成1220亿美元融资", "OpenAI获得1220亿美元新一轮融资，创AI行业最大融资纪录。资金用于全球扩展前沿AI、投资新一代算力、回应ChatGPT/Codex/企业AI需求", "OpenAI", "2026-05-10", "⭐⭐⭐⭐⭐ AI军备竞赛白热化，OpenAI资金储备远超竞争对手"],
    [3, "OpenAI $500神秘硬件设备泄露", "OpenAI正在开发约500美元的硬件设备，被认为是AI原生硬件入口，与苹果/字节的AI手机战略形成竞争", "OpenAI", "2026-05-11", "⭐⭐⭐⭐ AI从软件走向硬件，可能重塑人机交互方式"],
    [4, "Claude Opus 4.7发布", "Anthropic发布旗舰模型Claude Opus 4.7，文档推理能力大幅提升，OfficeQA Pro分数刷新纪录", "Anthropic", "2026-05-04", "⭐⭐⭐⭐ 企业AI竞争加剧，文档理解领域超越GPT-5.5"],
    [5, "Anthropic与SpaceX达成5GW算力协议", "Anthropic与SpaceX签署5GW算力协议，同时与Amazon达成近1GW新产能，Claude Code使用上限大幅提升", "Anthropic/SpaceX/Amazon", "2026-05-06", "⭐⭐⭐⭐ 算力成为AI公司核心竞争壁垒"],
    [6, "Anthropic+Blackstone成立企业AI公司", "Anthropic与Blackstone、Goldman Sachs联合成立新企业AI服务公司，瞄准金融/法律等高价值垂直领域", "Anthropic/Blackstone/Goldman", "2026-05-06", "⭐⭐⭐⭐ AI+金融深度融合，华尔街直接下场"],
    [7, "NVIDIA Rubin架构发布", "NVIDIA在CES发布下一代AI计算平台Rubin，Jensen Huang称将推出'震惊世界'的新芯片", "NVIDIA", "2026-05-05", "⭐⭐⭐⭐⭐ AI算力基础设施代际升级，训练成本持续下降"],
    [8, "NVIDIA面临自研芯片威胁", "Google/Amazon/Microsoft等科技巨头纷纷自研AI芯片，NVIDIA股价承压，成为半导体指数中表现最差成分股", "多家", "2026-05-06", "⭐⭐⭐⭐ AI芯片市场从垄断走向多元，长期利好算力成本下降"],
    [9, "Google DeepMind发布Deep Research Max", "基于Gemini 3.1 Pro的自主AI研究代理，支持MCP协议，可自主执行深度研究任务", "Google DeepMind", "2026-05-08", "⭐⭐⭐⭐ 自主AI Agent从概念到产品，MCP协议与ATEX兼容"],
    [10, "Google发布新AI芯片挑战NVIDIA", "Google推出新一代自研AI芯片，专为推理计算优化，与TPU v6架构结合", "Google", "2026-05-06", "⭐⭐⭐ 自研芯片趋势加速，推理芯片市场成为新战场"],
    [11, "AI融资四连跳破300亿大关", "中国AI赛道融资额连续四季度增长突破300亿元。NVIDIA承诺400亿美元AI投资，Intel股价暴涨490%", "多家", "2026-05-09", "⭐⭐⭐⭐ 资本持续涌入AI赛道，中国市场尤其活跃"],
    [12, "中国大厂AI入口争夺战", "字节豆包周活1.55亿+豆包手机，阿里千问生态接入，腾讯加入AI入口竞争", "字节/阿里/腾讯/百度", "2026-05-10", "⭐⭐⭐⭐ AI消费入口格局初现，Agent服务市场是下一层竞争"],
    [13, "智源发布2026十大AI技术趋势", "认知/形态/基建三重变革：人形机器人从Demo转向真实场景，开源模型性能追平闭源", "智源研究院", "2026-05-05", "⭐⭐⭐ Agent经济方向确认，开源生态持续壮大"],
    [14, "OpenAI推出Trusted Contact安全功能", "ChatGPT新增Trusted Contact可选安全功能，允许成年人指定信任联系人", "OpenAI", "2026-05-08", "⭐⭐ AI安全从自律走向产品化"],
    [15, "Cursor vs GitHub Copilot 2026对比", "AI编程助手竞争白热化：Cursor深度代码理解 vs Copilot生态集成。Warp开源AI终端+Cursor开放编码引擎", "Cursor/GitHub", "2026-05-10", "⭐⭐⭐ AI编程工具分化，深度vs广度两条路线"],
    [16, "Intel股价暴涨490%", "Intel股价从低点暴涨490%，市场押注芯片行业复苏和代工业务转型", "Intel", "2026-05-10", "⭐⭐⭐ 半导体行业全面复苏，AI算力需求持续增长"],
]

headers1 = ["序号", "新技术/事件", "主要功能或内容（外行人能理解）", "开发方/来源", "时间", "对行业影响"]
next_row = write_table(ws1, 3, headers1, news_data)

# Set row heights for data rows
for r in range(4, 4 + len(news_data)):
    ws1.row_dimensions[r].height = 52

# Column widths
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 55
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 42

# ============================================================
# Sheet2: 服务板块运营数据
# ============================================================
ws2 = wb.create_sheet("服务板块运营数据")

# Title
ws2.merge_cells('A1:F1')
t2 = ws2['A1']
t2.value = "ATEX服务板块运营数据 — 2026年5月13日"
t2.font = TITLE_FONT
t2.fill = DARK_FILL
t2.alignment = CENTER_WRAP
ws2.row_dimensions[1].height = 36

# Section 1: 服务总览
write_section_title(ws2, 3, 1, 4, "▎服务总览")
overview = [
    ["服务总数", 23, "活跃服务", 23],
    ["累计销量", 5, "累计收入(ATEX)", 35],
    ["分类数", 10, "新增服务(近期)", 5],
]
next_r = write_table(ws2, 4, ["指标", "数值", "指标", "数值"], overview)

# Section 2: 分类分布
write_section_title(ws2, 9, 1, 4, "▎各分类服务分布")
cat_data = [
    ["工具调用", 10, 7, 35],
    ["AI基础设施", 4, 0, 0],
    ["运营分析", 4, 0, 0],
    ["金融", 2, 0, 0],
    ["信息情报", 2, 0, 0],
    ["安全", 1, 0, 0],
    ["合规", 1, 0, 0],
    ["通信", 1, 0, 0],
    ["内容", 1, 0, 0],
    ["平台开发", 1, 0, 0],
]
next_r = write_table(ws2, 10, ["分类", "服务数", "累计销量", "累计收入(ATEX)"], cat_data)

# Section 3: 服务更新记录
write_section_title(ws2, 22, 1, 4, "▎昨日服务更新记录（5月11日20:00执行）")
update_data = [
    ["svc_001", "多模型路由与成本优化", "描述更新", "新增中国模型价格战信息，成本降低预期从30-50%提升至50-80%"],
    ["svc_002", "AI安全攻防服务", "描述更新", "新增军工级安全评估能力（五角大楼AI合同标准）"],
    ["svc_003", "AI法律合规与政策追踪", "描述更新", "新增GPT-5.5 Instant法律准确性突破、白宫国家AI政策框架"],
    ["svc_004", "实时语音翻译", "描述更新", "新增GPT-5级推理驱动说明、情感理解能力"],
    ["svc_005", "金融投研分析", "描述更新", "新增GPT-5.5 Instant金融幻觉显著减少"],
    ["svc_006", "内容质量审核", "描述更新", "新增AI生成内容专项优化（深度伪造检测）"],
    ["svc_017", "视频理解与生成", "描述更新", "新增Seedance 2.0和Muse Spark技术栈说明"],
]
next_r = write_table(ws2, 23, ["服务ID", "服务名称", "更新类型", "更新详情"], update_data)
for r in range(24, 24 + len(update_data)):
    ws2.row_dimensions[r].height = 32

# Section 4: 新增服务
write_section_title(ws2, 32, 1, 6, "▎近期新增服务")
new_svc = [
    ["svc_021", "企业Agent治理平台", "AI基础设施", "500/月", "Agent治理需求爆发", "强"],
    ["svc_023", "AI终端与编码引擎集成", "工具调用", "50/次", "Warp开源AI终端+Cursor开放编码引擎", "强"],
    ["svc_024", "AI创业融资顾问", "金融", "500/次", "Q1 2026全球VC $3000亿（AI占81%）", "强"],
    ["svc_025", "长文档深度分析", "工具调用", "50/份", "Gemini 3.1 Ultra 200万token上下文", "中强"],
    ["svc_026", "多模态内容理解", "工具调用", "30/条", "Qwen3.5原生VLM+Seedance 2.0", "强"],
]
next_r = write_table(ws2, 33, ["服务ID", "名称", "分类", "价格(ATEX)", "市场信号", "信号强度"], new_svc)

# Section 5: 待办行动项
write_section_title(ws2, 40, 1, 3, "▎待办行动项")
actions_svc = [
    ["P0", "多模型路由服务验证：需实际测试中国模型（豆包/文心/通义）API接入和成本对比", "待执行"],
    ["P0", "法律合规服务MVP：GPT-5.5 Instant法律准确性突破是关键拐点，优先构建法规数据库", "待执行"],
    ["P1", "企业Agent治理平台定价策略：500 ATEX/月偏低，考虑分级定价", "待决策"],
    ["P2", "AI芯片供应链分析Agent（储备方向）：NVIDIA面临自研芯片威胁+华为昇腾崛起", "储备"],
    ["P2", "AI设计工具集成Agent（储备方向）：Claude Connectors生态扩展中", "储备"],
]
next_r = write_table(ws2, 41, ["优先级", "行动项", "状态"], actions_svc)
for r in range(42, 42 + len(actions_svc)):
    ws2.row_dimensions[r].height = 30

# Column widths
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 50
ws2.column_dimensions['E'].width = 35
ws2.column_dimensions['F'].width = 12

# ============================================================
# Sheet3: 交易平台运营数据
# ============================================================
ws3 = wb.create_sheet("交易平台运营数据")

# Title
ws3.merge_cells('A1:F1')
t3 = ws3['A1']
t3.value = "ATEX交易平台运营数据 — 2026年5月13日"
t3.font = TITLE_FONT
t3.fill = DARK_FILL
t3.alignment = CENTER_WRAP
ws3.row_dimensions[1].height = 36

# Section 1: Token交易
write_section_title(ws3, 3, 1, 3, "▎Token交易数据")
token_data = [
    ["当日交易量", "15.0 ATEX", "流动性不足"],
    ["当日交易笔数", "1", ""],
    ["最新成交价", "1.5 ATEX", ""],
    ["当日最高/最低价", "1.5 / 1.5 ATEX", ""],
    ["买卖价差", "0.2 (12.5%)", "价差偏大"],
    ["买单深度", "1,400.0 ATEX (1笔)", ""],
    ["卖单深度", "1,602.0 ATEX (2笔)", ""],
    ["历史总交易笔数", "2", ""],
    ["历史总佣金", "1.1 ATEX", ""],
]
next_r = write_table(ws3, 4, ["指标", "数值", "说明"], token_data)

# Section 2: 服务交易
write_section_title(ws3, 15, 1, 3, "▎服务交易数据")
svc_trade = [
    ["当日服务购买笔数", "4", "Web搜索与深度阅读最受欢迎"],
    ["热门服务", "svc_012 Web搜索与深度阅读", "4笔交易，5 ATEX/次"],
    ["服务交易佣金(taker)", "0.25 ATEX/笔", ""],
    ["服务交易佣金(maker)", "0.15 ATEX/笔", ""],
    ["工具调用分类销量", "7笔 / 35.0 ATEX", "唯一有销量的分类"],
]
next_r = write_table(ws3, 16, ["指标", "数值", "说明"], svc_trade)

# Section 3: 用户数据
write_section_title(ws3, 23, 1, 3, "▎用户数据")
user_data = [
    ["总账户数", "10", ""],
    ["当日新增", "4", "增长显著"],
    ["活跃账户", "6", "60%活跃率"],
    ["平台余额", "1,005,025.1 ATEX", "含冻结2,400"],
    ["agent_a余额", "9,971.25 ATEX", "冻结75"],
    ["agent_b余额", "5,027.3 ATEX", ""],
    ["疑似测试账户", "e2e_test", "低余额，需关注"],
]
next_r = write_table(ws3, 24, ["指标", "数值", "说明"], user_data)

# Section 4: 财务数据
write_section_title(ws3, 33, 1, 3, "▎财务数据")
finance_data = [
    ["Token交易佣金(当日)", "≈0.15 ATEX", "1笔交易"],
    ["服务交易佣金(当日)", "≈1.6 ATEX", "4笔交易"],
    ["累计总佣金(Token)", "1.1 ATEX", ""],
    ["owner余额", "0 ATEX", "佣金尚未结算"],
]
next_r = write_table(ws3, 34, ["指标", "数值", "说明"], finance_data)

# Section 5: 推广与安全
write_section_title(ws3, 40, 1, 3, "▎推广与安全")
promo_data = [
    ["GitHub仓库", "已发布", "0 Star，需主动推广"],
    ["GitHub Pages落地页", "已上线", "https://lm203688.github.io/atex/"],
    ["推广资料包", "已就绪", "promo/ATEX推广资料包.md"],
    ["Agent推广指令", "已就绪", "promo/Agent推广指令.md"],
    ["HTTP API", "阻塞项❌", "影响MCP/A2A注册和外部Agent接入"],
    ["异常事件", "1个", "e2e_test疑似测试账户"],
    ["风控触发", "无", ""],
]
next_r = write_table(ws3, 41, ["指标", "状态", "说明"], promo_data)

# Section 6: 待办行动项
write_section_title(ws3, 50, 1, 3, "▎待办行动项")
actions_plat = [
    ["P0", "HTTP API仍为阻塞项，影响MCP/A2A注册和外部Agent接入——最高优先级", "阻塞❌"],
    ["P0", "Token交易量仅15.0，流动性不足——需做市商机制或激励计划", "待执行"],
    ["P1", "GitHub 0 Star，需主动推广到开发者社区（HackerNews/Reddit/V2EX）", "待执行"],
    ["P1", "今日4个新账户注册，需跟进激活引导（首次交易引导流程）", "待执行"],
    ["P2", "Web搜索最受欢迎，可考虑增加相关服务（深度研究/行业报告等）", "待规划"],
    ["P2", "检测到1个疑似测试/低余额账户(e2e_test)，需清理或标记", "待处理"],
]
next_r = write_table(ws3, 51, ["优先级", "行动项", "状态"], actions_plat)
for r in range(52, 52 + len(actions_plat)):
    ws3.row_dimensions[r].height = 30

# Column widths
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15
ws3.column_dimensions['F'].width = 15

# Save
output_path = "/home/z/my-project/reports/ATEX综合日报_20260513.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")
