#!/usr/bin/env python3
"""ATEX综合日报生成器 - 2026-05-29"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式定义 ──
font_header = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
font_body = Font(name='Microsoft YaHei', size=10)
font_title = Font(name='Microsoft YaHei', bold=True, size=14, color='1F4E79')
font_subtitle = Font(name='Microsoft YaHei', bold=True, size=11, color='2E75B6')
font_action = Font(name='Microsoft YaHei', size=10, color='C00000', bold=True)
font_note = Font(name='Microsoft YaHei', size=9, color='808080', italic=True)

fill_header = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
fill_alt = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
fill_section = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
fill_warn = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_ok = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

def style_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_header
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_body
        cell.alignment = align_left
        cell.border = thin_border
        if alt:
            cell.fill = fill_alt

def style_section_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    cell.fill = fill_section
    cell.alignment = align_left
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# Sheet1: 全球AI技术动态
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "全球AI技术动态"

# 标题
ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value="ATEX综合日报 - 全球AI技术动态").font = font_title
ws1.cell(row=1, column=1).alignment = align_center
ws1.merge_cells('A2:F2')
ws1.cell(row=2, column=1, value="2026年5月29日 | 数据来源：OpenAI/Google/Anthropic/Meta/NVIDIA/Microsoft/主流科技媒体").font = font_note
ws1.cell(row=2, column=1).alignment = align_center

# 表头
headers = ["序号", "新技术/事件名称", "主要功能/内容", "开发方/来源", "时间", "对行业影响"]
row = 4
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 6)

# AI动态数据（基于14组搜索结果整理）
ai_news = [
    [1, "GPT-5.5 / GPT-5.5 Pro", "OpenAI最新旗舰模型，推理+编码+Agent工作流全面升级，被Gartner评为企业编码Agent领导者", "OpenAI", "2026-04/05", "重新定义AI能力上限，企业编码Agent赛道领导者认证，API已开放"],
    [2, "Gemini 3.5 Flash + Gemini Omni", "前沿智能+行动能力，Gemini for Science科学发现工具集，Google I/O 2026发布", "Google DeepMind", "2026-05", "AI从对话走向行动，科学发现专用AI工具开创研究新范式"],
    [3, "Claude Opus 4.7 + Claude Design", "Anthropic最强模型GA，新增Claude Design设计产品，宣布永久无广告策略，米兰办公室", "Anthropic", "2026-05-04/27", "AI设计工具新赛道，无广告承诺差异化定位，欧洲市场扩张"],
    [4, "Llama 4 Scout/Maverick", "首个开源原生多模态模型，超长上下文支持，Llama 4.5即将发布", "Meta AI", "2026-Q1/Q2", "开源模型多模态里程碑，与闭源模型差距进一步缩小"],
    [5, "NVIDIA Rubin平台", "6款新芯片+AI超级计算机，CES 2026发布，但2026年无新游戏GPU（内存短缺）", "NVIDIA", "2026-01", "AI算力全面转向数据中心，70% DRAM分配给AI，游戏GPU让位"],
    [6, "Microsoft 365 Copilot重大更新", "Copilot Notebooks支持Pages创建编辑，3月AI组织架构重组，5月新功能发布", "Microsoft", "2026-03/05", "企业AI助手从工具升级为工作流核心，组织架构调整显示战略重心"],
    [7, "AI Agent协议生态四强并立", "MCP/A2A/ACP/UCP四大协议各有定位，MCP工具调用+A2A Agent通信+ACP新兴+UCP统一", "行业共识", "2026-Q1", "Agent互操作标准加速形成，ATEX svc_028协议适配器价值凸显"],
    [8, "大模型集体涨价潮", "海外主流订阅$8-20/月，国内大模型从免费转向收费，豆包/文心/通义均调整定价", "全行业", "2026-05", "免费时代结束，AI服务商业化加速，ATEX Token交易模式更具吸引力"],
    [9, "AI编程工具三强争霸", "Cursor 3/Claude Code/GitHub Copilot激烈竞争，开源Coding Agent崛起", "Cursor/Anthropic/GitHub", "2026-H1", "编程范式根本性变革，AI编码从辅助到主导，企业采用率飙升"],
    [10, "AI视频生成四强格局", "Sora 2/Veo 3.1/Kling 3.0/Runway Gen-4.5，4K+音频+API全面竞争", "OpenAI/Google/快手/Runway", "2026-H1", "视频生成从概念到生产级，价格下降至$0.5/5秒，ATEX svc_017受益"],
    [11, "智谱入选《时代》全球AI十强", "唯一上榜中国独立大模型企业，字节/阿里同期入选", "智谱AI/字节/阿里", "2026-04", "中国AI全球影响力提升，独立大模型公司获国际认可"],
    [12, "字节跳动AI投入1600亿", "2026年AI算力采购850亿，Doubao-Seed-2.0发布，文心5.0超稀疏MoE", "字节/百度", "2026", "中国AI军备竞赛白热化，算力投入超全球多数国家GDP"],
    [13, "xAI Grok接入Kilo Code", "Grok Build CLI支持内联diff和计划查看器，编程Agent新入口", "xAI", "2026-05-27", "xAI从对话AI向编程Agent扩展，与Cursor/Copilot直接竞争"],
    [14, "AI独角兽69家总估值6380亿美元", "14家超级独角兽贡献近半估值，硬科技占比超75%，平均4.5年晋级", "全球VC", "2026-Q1", "AI投资泡沫与价值并存，超级独角兽集中度提高，NVIDIA市值$4.8T"],
    [15, "Anthropic Economic Index报告", "Claude使用行为经济学分析，学习曲线研究揭示AI采用模式", "Anthropic", "2026-05", "AI经济影响量化研究，为Agent经济模型提供数据支撑"],
]

for idx, item in enumerate(ai_news):
    r = row + 1 + idx
    for c, val in enumerate(item, 1):
        ws1.cell(row=r, column=c, value=val)
    style_data_row(ws1, r, 6, alt=(idx % 2 == 1))

# 列宽
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 55
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 45

# ═══════════════════════════════════════════════════════════════
# Sheet2: 服务板块运营数据
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("服务板块运营数据")

ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value="ATEX综合日报 - 服务板块运营数据").font = font_title
ws2.cell(row=1, column=1).alignment = align_center
ws2.merge_cells('A2:F2')
ws2.cell(row=2, column=1, value="2026年5月29日 | 数据来源：daily_service_tracking.json + services.json").font = font_note
ws2.cell(row=2, column=1).alignment = align_center

r = 4
# ── 服务总览 ──
style_section_row(ws2, r, 6, "📊 服务总览")
r += 1
overview = [
    ["服务总数", 40, "活跃服务", 0, "暂停服务", 0],
    ["有销量服务", 13, "零销量服务", 27, "累计总销量", 26],
    ["累计总收入", "0 ATEX", "昨日新增", 0, "昨日下架", 0],
]
for item in overview:
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(r % 2 == 0))
    r += 1

r += 1
# ── 各分类服务数 ──
style_section_row(ws2, r, 6, "📁 各分类服务统计")
r += 1
cat_headers = ["分类", "服务数", "累计销量", "累计收入(ATEX)", "占比", "状态"]
for i, h in enumerate(cat_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)
r += 1

categories = [
    ["AI基础设施", 11, 8, 0, "27.5%", "核心赛道"],
    ["工具调用", 11, 11, 0, "27.5%", "销量最高"],
    ["运营分析", 4, 5, 0, "10.0%", "增长中"],
    ["金融", 4, 0, 0, "10.0%", "待激活"],
    ["信息情报", 4, 0, 0, "10.0%", "待激活"],
    ["安全", 2, 0, 0, "5.0%", "待激活"],
    ["内容", 2, 0, 0, "5.0%", "待激活"],
    ["合规", 1, 1, 0, "2.5%", "低频高值"],
    ["通信", 1, 1, 0, "2.5%", "低频高值"],
]
for idx, item in enumerate(categories):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 销量排行TOP10 ──
style_section_row(ws2, r, 6, "🏆 服务销量排行 TOP10")
r += 1
rank_headers = ["排名", "服务ID", "服务名称", "分类", "单价(ATEX)", "累计销量"]
for i, h in enumerate(rank_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)
r += 1

top_sold = [
    [1, "svc_012", "Web搜索与深度阅读", "工具调用", 5, 7],
    [2, "svc_015", "AI图像生成与编辑", "工具调用", 20, 3],
    [3, "svc_032", "Agent Marketplace上架服务", "运营分析", 60, 3],
    [4, "svc_001", "多模型路由与成本优化", "AI基础设施", 10, 2],
    [5, "svc_028", "Agent协议适配器", "AI基础设施", 80, 2],
    [6, "svc_030", "Agent算力成本优化", "AI基础设施", 60, 2],
    [7, "svc_003", "AI法律合规与政策追踪", "合规", 500, 1],
    [8, "svc_004", "实时语音翻译", "通信", 20, 1],
    [9, "svc_016", "语音合成与识别", "工具调用", 10, 1],
    [10, "svc_018", "运营数据分析与报告", "运营分析", 30, 1],
]
for idx, item in enumerate(top_sold):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 收入排行TOP5 ──
style_section_row(ws2, r, 6, "💰 服务收入排行 TOP5（按单价×销量估算）")
r += 1
rev_headers = ["排名", "服务ID", "服务名称", "单价(ATEX)", "销量", "估算收入(ATEX)"]
for i, h in enumerate(rev_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)
r += 1

top_rev = [
    [1, "svc_003", "AI法律合规与政策追踪", 500, 1, 500],
    [2, "svc_032", "Agent Marketplace上架服务", 60, 3, 180],
    [3, "svc_030", "Agent算力成本优化", 60, 2, 120],
    [4, "svc_028", "Agent协议适配器", 80, 2, 160],
    [5, "svc_015", "AI图像生成与编辑", 20, 3, 60],
]
for idx, item in enumerate(top_rev):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 昨日服务更新 ──
style_section_row(ws2, r, 6, "🔄 近期服务更新记录")
r += 1
upd_headers = ["服务ID", "服务名称", "更新类型", "更新详情", "", ""]
for i, h in enumerate(upd_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)
r += 1

updates = [
    ["svc_028", "Agent协议适配器", "描述更新", "新增ACP第三大协议支持，新增Microsoft Agent Framework 1.0生产级需求引用", "", ""],
    ["svc_022", "开源AI模型集成", "描述更新", "新增Llama 4/Kimi K2.6/Mistral Medium 3.5三大前沿开源模型", "", ""],
    ["svc_001", "多模型路由与成本优化", "描述更新", "新增Llama 4和Kimi K2.6路由支持", "", ""],
    ["svc_029", "空间智能与3D内容生成", "描述更新", "新增World Labs/SpAItial/Niantic Spatial/MetaEarth3D等世界模型引用", "", ""],
    ["svc_021", "企业Agent治理平台", "描述更新", "新增Microsoft Agent Framework 1.0对标，补充多协议互操作安全管控", "", ""],
]
for idx, item in enumerate(updates):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 新增服务 ──
style_section_row(ws2, r, 6, "🆕 新增服务")
r += 1
new_headers = ["服务ID", "名称", "分类", "单价(ATEX)", "市场信号", "信号强度"]
for i, h in enumerate(new_headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)
r += 1

new_svcs = [
    ["svc_032", "Agent Marketplace上架服务", "运营分析", 60, "8大Agent市场并存，7.84B→52.62B(2030)", "极强"],
    ["svc_033", "AI Agent测试与评估", "运营分析", 40, "Agent数量爆发，benchlm.ai等基准平台兴起", "强"],
]
for idx, item in enumerate(new_svcs):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 待办行动项 ──
style_section_row(ws2, r, 6, "📋 待办行动项")
r += 1
actions = [
    "1. Agent Marketplace上架服务MVP：优先对接Claude Skills Store和MCP Hubs",
    "2. AI Agent测试与评估：建立标准化测试框架，参考benchlm.ai基准",
    "3. svc_021企业Agent治理平台定价策略：500 ATEX/月偏低，考虑分级定价",
    "4. ACP协议跟踪：作为第三大Agent协议快速崛起，持续关注生态成熟度",
    "5. 开源模型价格战跟踪：持续更新svc_001和svc_022的模型列表和成本数据",
    "6. 空间智能赛道深化：World Labs已发布产品，svc_029可考虑增加3D生成API对接",
]
for item in actions:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws2.cell(row=r, column=1, value=item).font = font_action
    ws2.cell(row=r, column=1).alignment = align_left
    r += 1

# 列宽
ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 14

# ═══════════════════════════════════════════════════════════════
# Sheet3: 交易平台运营数据
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("交易平台运营数据")

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value="ATEX综合日报 - 交易平台运营数据").font = font_title
ws3.cell(row=1, column=1).alignment = align_center
ws3.merge_cells('A2:F2')
ws3.cell(row=2, column=1, value="2026年5月29日 | 数据来源：daily_platform_ops.json + bootstrap_report.json + atex.py引擎").font = font_note
ws3.cell(row=2, column=1).alignment = align_center

r = 4
# ── Token交易数据 ──
style_section_row(ws3, r, 6, "💱 Token交易数据（ATEX）")
r += 1
token_data = [
    ["指标", "数值", "指标", "数值", "指标", "数值"],
    ["当日交易量", 0, "当日交易笔数", 0, "最新成交价", "1.5 ATEX"],
    ["当日最高价", "N/A", "当日最低价", "N/A", "历史总交易笔数", 1],
    ["挂单买盘", "1.0×10", "挂单卖盘", "1.5×50", "订单簿深度", "2档"],
    ["佣金费率", "Maker 3%/Taker 5%", "", "", "", ""],
]
for idx, item in enumerate(token_data):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
    r += 1

r += 1
# ── 历史成交 ──
style_section_row(ws3, r, 6, "📜 历史成交记录")
r += 1
trade_headers = ["成交ID", "买方", "卖方", "价格", "数量", "时间"]
for i, h in enumerate(trade_headers, 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, 6)
r += 1

trades = [
    ["46dffe0e8fae", "trader_b", "trader_a", 1.5, 50, "2026-05-18 07:35:22"],
]
for idx, item in enumerate(trades):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, 6)
    r += 1

r += 1
# ── 服务交易数据 ──
style_section_row(ws3, r, 6, "🛒 服务交易数据")
r += 1
svc_trade = [
    ["指标", "数值", "指标", "数值", "指标", "数值"],
    ["服务总数", 40, "服务购买总笔数", 26, "有销量服务", 13],
    ["热门服务#1", "Web搜索与深度阅读(7次)", "热门服务#2", "AI图像生成(3次)", "热门服务#3", "Marketplace上架(3次)"],
    ["虚拟买家总数", 16, "虚拟购买笔数", 2, "虚拟购买金额", "160 ATEX"],
]
for idx, item in enumerate(svc_trade):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
    r += 1

r += 1
# ── Provider收入排行 ──
style_section_row(ws3, r, 6, "🏢 Provider收入排行")
r += 1
prov_headers = ["排名", "Provider", "累计收入(ATEX)", "服务数", "主要服务", ""]
for i, h in enumerate(prov_headers, 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, 6)
r += 1

providers = [
    [1, "ai_infra_provider", 120, "-", "Agent算力成本优化", ""],
    [2, "content_provider", 40, "-", "AI图像生成与编辑", ""],
    [3, "platform", "（平台服务）", "40", "全品类", ""],
]
for idx, item in enumerate(providers):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, 6, alt=(idx % 2 == 1))
    r += 1

r += 1
# ── 用户数据 ──
style_section_row(ws3, r, 6, "👥 用户数据")
r += 1
user_data = [
    ["指标", "数值", "指标", "数值", "指标", "数值"],
    ["总账户数", 44, "SaaS注册用户", 1, "虚拟买家", 16],
    ["新增用户(当日)", 0, "活跃用户(当日)", 0, "虚拟vs真实占比", "虚拟为主"],
]
for idx, item in enumerate(user_data):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
    r += 1

r += 1
# ── 佣金数据 ──
style_section_row(ws3, r, 6, "💸 佣金数据")
r += 1
comm_data = [
    ["指标", "数值", "指标", "数值", "指标", "数值"],
    ["当日佣金", 0, "累计佣金", "20.80 ATEX", "佣金费率", "Maker 3%/Taker 5%"],
    ["历史最大单笔佣金(Taker)", 3.75, "历史最大单笔佣金(Maker)", 2.25, "佣金来源", "Token交易"],
]
for idx, item in enumerate(comm_data):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
    r += 1

r += 1
# ── 推广效果 ──
style_section_row(ws3, r, 6, "📢 推广效果")
r += 1
promo_data = [
    ["渠道", "状态", "效果", "", "", ""],
    ["GitHub落地页", "✅ 已启用", "https://lm203688.github.io/atex/", "", "", ""],
    ["Agent推广指令", "✅ 已发布", "promo/Agent推广指令.md", "", "", ""],
    ["推广资料包", "✅ 已发布", "promo/ATEX推广资料包.md", "", "", ""],
    ["外部Provider邀请", "⏳ 待执行", "5类Provider邀请计划（AI模型/数据/安全/开发/内容）", "", "", ""],
]
for idx, item in enumerate(promo_data):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
    r += 1

r += 1
# ── 安全事件 ──
style_section_row(ws3, r, 6, "🔒 安全事件")
r += 1
sec_data = [
    ["指标", "状态", "详情", "", "", ""],
    ["当日安全事件", 0, "无异常", "", "", ""],
    ["ECS API状态", "⚠️ 不可达", "Connection timed out (10s)，DNS解析失败", "", "", ""],
    ["本地引擎", "✅ 正常", "v5.10运行中", "", "", ""],
    ["ECS宕机检测", "⚠️ 已检测", "ECS服务不可用，数据来自本地引擎", "", "", ""],
]
for idx, item in enumerate(sec_data):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    if idx == 0:
        style_header_row(ws3, r, 6)
    else:
        style_data_row(ws3, r, 6, alt=(idx % 2 == 0))
        # 高亮警告行
        if "⚠️" in str(item[1]):
            for c in range(1, 7):
                ws3.cell(row=r, column=c).fill = fill_warn
        elif "✅" in str(item[1]):
            for c in range(1, 7):
                ws3.cell(row=r, column=c).fill = fill_ok
    r += 1

r += 1
# ── 待办行动项 ──
style_section_row(ws3, r, 6, "📋 待办行动项")
r += 1
actions3 = [
    "1. 【紧急】ECS API不可达：排查150.158.119.19:8420连接问题，检查tunnel DNS解析",
    "2. Token交易量低迷：考虑做市商策略或引入流动性激励",
    "3. 服务交易零收入：确认revenue计算逻辑，26笔购买但收入为0需排查",
    "4. 虚拟vs真实交易占比失衡：加速外部Provider入驻，引入真实交易",
    "5. 推广执行：5类Provider邀请计划需启动，优先AI模型提供商",
    "6. 佣金结构优化：当前Maker 3%/Taker 5%偏高，考虑阶梯费率吸引做市商",
]
for item in actions3:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws3.cell(row=r, column=1, value=item).font = font_action
    ws3.cell(row=r, column=1).alignment = align_left
    r += 1

# 列宽
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 35
ws3.column_dimensions['F'].width = 14

# ── 保存 ──
output_path = "/home/z/my-project/reports/ATEX综合日报_20260529.xlsx"
wb.save(output_path)
print(f"✅ 报告已保存: {output_path}")
