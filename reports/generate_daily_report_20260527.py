#!/usr/bin/env python3
"""ATEX综合日报生成器 - 2026-05-27"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 样式定义 =====
font_header = lambda: Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
font_body = lambda: Font(name='微软雅黑', size=10)
font_title = lambda: Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
font_subtitle = lambda: Font(name='微软雅黑', bold=True, size=11, color='2E75B6')
font_highlight = lambda: Font(name='微软雅黑', bold=True, size=10, color='C00000')
font_note = lambda: Font(name='微软雅黑', size=9, color='808080')

fill_header = lambda: PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
fill_alt = lambda: PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
fill_white = lambda: PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
fill_yellow = lambda: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_green = lambda: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
fill_red = lambda: PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

def apply_header_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_header()
        cell.fill = fill_header()
        cell.alignment = align_center
        cell.border = thin_border

def apply_body_row(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_body()
        cell.fill = fill_alt() if alt else fill_white()
        cell.alignment = align_left
        cell.border = thin_border

# ===== 创建工作簿 =====
wb = Workbook()

# ============================================================
# Sheet1: 全球AI技术动态
# ============================================================
ws1 = wb.active
ws1.title = "全球AI技术动态"

# 标题
ws1.merge_cells('A1:F1')
ws1['A1'] = '全球AI技术动态日报 | 2026-05-27'
ws1['A1'].font = font_title()
ws1['A1'].alignment = align_center

ws1.merge_cells('A2:F2')
ws1['A2'] = '数据来源：OpenAI/Google/Anthropic/Meta/NVIDIA/Microsoft/百度/阿里/腾讯/字节跳动官方发布 + The Verge/TechCrunch/机器之心/量子位/36氪等媒体'
ws1['A2'].font = font_note()
ws1['A2'].alignment = align_center

# 表头
headers1 = ['序号', '技术/产品名称', '主要功能与亮点', '开发方', '发布/更新时间', '对行业影响评估']
row = 4
for i, h in enumerate(headers1, 1):
    ws1.cell(row=row, column=i, value=h)
apply_header_row(ws1, row, 6)

# AI动态数据
ai_news = [
    [1, 'GPT-5.5 & GPT-5.5 Pro',
     'OpenAI最新旗舰模型，API已开放。指令遵循显著提升，编码/研究/数据分析跨工具能力增强。下一代推理效率优化，网络安全能力突破。ChatGPT新增richer context和goal mode，Codex更新浏览器改进。OpenAI向每家YC公司提供$2M token额度。',
     'OpenAI', '2026-04-23(API)/5月持续更新', '★★★★★ 旗舰模型迭代加速，API生态绑定YC系创业公司，企业AI应用成本进一步降低'],
    [2, 'Gemini 3.5 & 3.5 Flash',
     'Google I/O 2026发布。前沿智能+行动能力结合，Gemini 3.5 Flash已集成Google Search。Google Spark个人Agent发布。多模态搜索升级，File Search支持多模态。Cloud AI全面集成Gemini 3.5。',
     'Google DeepMind', '2026-05-19(I/O)', '★★★★★ Google将Gemini深度嵌入Search/Cloud/Android全栈，AI Agent从对话走向行动'],
    [3, 'Claude Opus 4.7 & Claude Design',
     'Anthropic最新Opus模型：编码/Agent/视觉/多步骤任务全面增强。Claude Design（Anthropic Labs）：AI设计工具。Claude Code改进后台会话和代码审查。提高使用限额，与SpaceX达成算力合作。模型弃用通知：Claude Sonnet 4等旧模型将下线。',
     'Anthropic', '2026-04-16(Opus)/4-17(Design)/5月更新', '★★★★☆ Anthropic从安全研究走向产品化，Design工具切入创意市场，SpaceX算力合作值得关注'],
    [4, 'Grok Build CLI',
     'xAI发布Grok Build命令行工具（早期Beta）：终端内AI编码，支持inline diffs和plan viewer。Grok Voice语音Agent（亚秒延迟）。Grok Imagine图像/视频生成。Grok 5传闻：xAI声称10%概率实现AGI。',
     'xAI', '2026-05-25(Build)', '★★★☆☆ xAI从聊天扩展到编码/语音/图像全栈，但Grok 5 AGI声明争议大'],
    [5, 'NVIDIA Rubin架构',
     'NVIDIA下一代AI芯片Rubin发布：6款新芯片，推理收入机会预测从$500B大幅上调。2026年游戏GPU可能跳过更新，全力转向AI芯片。70% DRAM产能分配给AI。CPU在AI推理中地位上升。',
     'NVIDIA', '2026年持续', '★★★★★ AI芯片需求持续爆发，推理市场成新增长极，游戏GPU让位AI是标志性事件'],
    [6, 'Meta Llama 4 / Avocado延迟',
     'Llama 4 Scout/Maverick发布（原生多模态开放权重）。但下一代"Avocado"模型延迟发布（原定2025年底→推迟至5月+），性能不达标，可能转向闭源。Meta AI探索直接变现。',
     'Meta', '2026年5月', '★★★☆☆ 开源Llama 4发布但旗舰模型延迟，Meta在开源vs闭源间摇摆，闭源倾向令人担忧'],
    [7, 'Microsoft 365 Copilot重大更新',
     '5月Copilot更新：文本格式支持、多模态交互（语音/图像）、Windows深度集成。3月Copilot部门重组（Satya Nadella宣布）。Copilot Key被放弃。独立报告显示Copilot采用率仍低于预期。',
     'Microsoft', '2026年5月', '★★★☆☆ Copilot功能持续迭代但采用率存疑，组织重组反映战略调整压力'],
    [8, '中国大模型三国杀：阿里/字节/腾讯',
     '2026年4月密集发布潮：阿里千问APP支持AI点外卖/订机票（生态闭环）；字节豆包2.0+Seedance 2.0+Seedream 5.0三模型礼包，豆包月活破亿；腾讯马化腾罕见公开谈AI战略。三巨头2026年AI capex合计超600亿美元。',
     '阿里/字节/腾讯', '2026年1-5月', '★★★★★ 中国AI入口争夺战白热化，字节倒逼大厂加速，AI从模型竞争转向生态/入口竞争'],
    [9, 'AI Agent协议生态：MCP/A2A/ACP',
     '三大协议并存互补：MCP（工具调用+上下文管理）、A2A（Agent间协作+任务分发）、ACP（新兴第三标准快速获生态支持）。MCP安全风险受关注（Cyber Strategy Institute报告）。Orca Security指出协议内存管理挑战。Boomi/GetStream等发布集成指南。',
     'Anthropic/Google/社区', '2026年5月', '★★★★☆ Agent协议从概念走向生产，MCP安全审计成新赛道，多协议互操作是核心挑战'],
    [10, 'AI视频生成六强争霸',
     'Sora 2（电影级画质但4月26日停售）、Runway Gen-4/4.5（专业控制力最强）、Kling 2.0（性价比之王）、Seedance 2.0（API集成领先）、WAN 2.2、Luma。Sora停售后竞品用户激增。6个AI视频模型在TikTok爆发。',
     'OpenAI/Runway/Kuaishou/ByteDance等', '2026年5月', '★★★★☆ Sora停售标志AI视频从技术验证进入商业竞争，API集成和成本效率成关键'],
    [11, 'AI编码Agent五强对比',
     'Claude Code vs Cursor vs GitHub Copilot vs Devin vs Windsurf。Cursor获Forbes报道"AI编码霸权之战"。Devin定位自主编码但口碑分化。5月22日对比评测：Claude Code代码审查强，Cursor生态最广。',
     'Anthropic/Cursor/GitHub/Cognition等', '2026年5月', '★★★★☆ AI编码从辅助走向自主，Cursor增长最快但竞争加剧，Agent编码是确定性趋势'],
    [12, 'EU AI Act重大更新',
     '5月7日欧盟达成临时协议：推迟部分合规截止日期、定向简化要求、新增禁止条款。8月全面执行，罚款最高全球营收7%。中美欧三套监管模式分化：EU预防式、US市场驱动、China 5月新发AI Agent框架。',
     '欧盟委员会', '2026-05-07', '★★★★☆ 全球AI监管三极分化加剧，EU简化调整反映执行压力，中国Agent框架值得关注'],
    [13, 'AI融资创纪录：Q1 2026',
     'Q1 2026全球VC融资创纪录，AI占81%（约$244B）。AI独占全球VC三分之一。5月14日单日10笔融资超$120M。AI创业融资趋势：个人储蓄→种子轮→A/B轮加速。',
     'Crunchbase/MUFG/多机构', '2026 Q1', '★★★★★ AI融资泡沫还是黄金时代？81%占比史无前例，资金向头部集中'],
    [14, 'OpenAI模型弃用潮',
     'OpenAI宣布未来3-6个月将弃用一批旧模型。Claude Sonnet 4/Claude Opus 4也收到弃用通知（4月14日）。模型迭代加速，迁移成本成为企业新痛点。',
     'OpenAI/Anthropic', '2026年5月', '★★★☆☆ 模型生命周期缩短，企业需建立模型迁移策略，ATEX多模型路由服务价值凸显'],
]

for i, news in enumerate(ai_news):
    r = row + 1 + i
    for j, val in enumerate(news):
        ws1.cell(row=r, column=j+1, value=val)
    apply_body_row(ws1, r, 6, alt=(i % 2 == 1))
    # 影响评估列高亮
    impact_cell = ws1.cell(row=r, column=6)
    if '★★★★★' in str(impact_cell.value):
        impact_cell.font = font_highlight()

# 列宽
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 65
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 50

# svc_010描述更新
ws1.merge_cells(f'A{row+1+len(ai_news)+2}:F{row+1+len(ai_news)+2}')
svc010_row = row + 1 + len(ai_news) + 2
ws1.cell(row=svc010_row, column=1, value='svc_010 AI信息情报收集 - 描述更新')
ws1.cell(row=svc010_row, column=1).font = font_subtitle()

ws1.merge_cells(f'A{svc010_row+1}:F{svc010_row+1}')
new_desc = ('全球AI技术动态实时追踪与深度分析。覆盖OpenAI/Google/Anthropic/Meta/百度/阿里/腾讯/字节等核心厂商，'
            'MCP/A2A/ACP等Agent协议，AI芯片/算力/融资/监管等关键赛道。每日更新，提供技术解读+行业影响评估+ATEX服务关联建议。'
            '最新动态：GPT-5.5 API开放+Codex升级、Gemini 3.5 I/O发布+Search集成、Claude Opus 4.7+Design工具、'
            'NVIDIA Rubin架构+推理收入上调、EU AI Act临时协议+三极监管分化、AI融资Q1创纪录占81%、'
            'AI视频六强争霸(Sora停售)、AI编码Agent五强对比、中国大模型入口争夺战白热化。')
ws1.cell(row=svc010_row+1, column=1, value=new_desc)
ws1.cell(row=svc010_row+1, column=1).font = font_body()
ws1.cell(row=svc010_row+1, column=1).alignment = align_left_top

# ============================================================
# Sheet2: 服务板块运营数据
# ============================================================
ws2 = wb.create_sheet("服务板块运营数据")

# 读取数据
with open('/home/z/my-project/token_exchange/data/daily_service_tracking.json') as f:
    svc_tracking = json.load(f)
with open('/home/z/my-project/services/services.json') as f:
    svc_data = json.load(f)

# 标题
ws2.merge_cells('A1:F1')
ws2['A1'] = 'ATEX服务板块运营数据 | 2026-05-27'
ws2['A1'].font = font_title()
ws2['A1'].alignment = align_center

# === 概览 ===
ws2.merge_cells('A3:F3')
ws2['A3'] = '一、服务总览'
ws2['A3'].font = font_subtitle()

overview_data = [
    ['指标', '数值'],
    ['服务总数（活跃）', f'{svc_tracking["total_active"]}'],
    ['服务总数（暂停）', f'{svc_tracking["total_paused"]}'],
    ['服务总数（合计）', f'{svc_tracking["services_checked"]}'],
    ['昨日服务更新', f'{len(svc_tracking["updates"])}项'],
    ['新增服务', f'{len(svc_tracking["new_services"])}项'],
    ['定价变更', f'{len(svc_tracking["pricing_changes"])}项'],
    ['待办行动项', f'{len(svc_tracking["action_items"])}项'],
]

for i, row_data in enumerate(overview_data):
    r = 4 + i
    for j, val in enumerate(row_data):
        ws2.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws2, r, 2)
    else:
        apply_body_row(ws2, r, 2, alt=(i % 2 == 0))

# === 分类统计 ===
cat_start = 4 + len(overview_data) + 1
ws2.merge_cells(f'A{cat_start}:F{cat_start}')
ws2.cell(row=cat_start, column=1, value='二、各分类服务数').font = font_subtitle()

categories = svc_tracking.get('service_summary', [])
# Aggregate by category
cat_counts = {}
for s in categories:
    cat = s['category']
    if cat not in cat_counts:
        cat_counts[cat] = {'count': 0, 'sold': 0, 'revenue': 0}
    cat_counts[cat]['count'] += 1
    cat_counts[cat]['sold'] += s.get('total_sold', 0)
    cat_counts[cat]['revenue'] += s.get('total_sold', 0) * s.get('price', 0)

cat_headers = ['分类', '服务数', '总销量', '总收入(ATEX)']
r = cat_start + 1
for j, h in enumerate(cat_headers):
    ws2.cell(row=r, column=j+1, value=h)
apply_header_row(ws2, r, 4)

sorted_cats = sorted(cat_counts.items(), key=lambda x: x[1]['revenue'], reverse=True)
for i, (cat, data) in enumerate(sorted_cats):
    r = cat_start + 2 + i
    ws2.cell(row=r, column=1, value=cat)
    ws2.cell(row=r, column=2, value=data['count'])
    ws2.cell(row=r, column=3, value=data['sold'])
    ws2.cell(row=r, column=4, value=data['revenue'])
    apply_body_row(ws2, r, 4, alt=(i % 2 == 1))

# === 销量排行 ===
sold_start = cat_start + 2 + len(sorted_cats) + 1
ws2.merge_cells(f'A{sold_start}:F{sold_start}')
ws2.cell(row=sold_start, column=1, value='三、服务销量排行（Top 10）').font = font_subtitle()

sold_headers = ['排名', '服务ID', '服务名称', '分类', '单价(ATEX)', '累计销量']
r = sold_start + 1
for j, h in enumerate(sold_headers):
    ws2.cell(row=r, column=j+1, value=h)
apply_header_row(ws2, r, 6)

sorted_by_sold = sorted(categories, key=lambda x: x.get('total_sold', 0), reverse=True)[:10]
for i, s in enumerate(sorted_by_sold):
    r = sold_start + 2 + i
    ws2.cell(row=r, column=1, value=i+1)
    ws2.cell(row=r, column=2, value=s['id'])
    ws2.cell(row=r, column=3, value=s['name'])
    ws2.cell(row=r, column=4, value=s['category'])
    ws2.cell(row=r, column=5, value=s['price'])
    ws2.cell(row=r, column=6, value=s.get('total_sold', 0))
    apply_body_row(ws2, r, 6, alt=(i % 2 == 1))

# === 收入排行 ===
rev_start = sold_start + 2 + len(sorted_by_sold) + 1
ws2.merge_cells(f'A{rev_start}:F{rev_start}')
ws2.cell(row=rev_start, column=1, value='四、服务收入排行（Top 10）').font = font_subtitle()

rev_headers = ['排名', '服务ID', '服务名称', '分类', '单价(ATEX)', '累计收入(ATEX)']
r = rev_start + 1
for j, h in enumerate(rev_headers):
    ws2.cell(row=r, column=j+1, value=h)
apply_header_row(ws2, r, 6)

sorted_by_rev = sorted(categories, key=lambda x: x.get('total_sold', 0) * x.get('price', 0), reverse=True)[:10]
for i, s in enumerate(sorted_by_rev):
    r = rev_start + 2 + i
    rev = s.get('total_sold', 0) * s.get('price', 0)
    ws2.cell(row=r, column=1, value=i+1)
    ws2.cell(row=r, column=2, value=s['id'])
    ws2.cell(row=r, column=3, value=s['name'])
    ws2.cell(row=r, column=4, value=s['category'])
    ws2.cell(row=r, column=5, value=s['price'])
    ws2.cell(row=r, column=6, value=rev)
    apply_body_row(ws2, r, 6, alt=(i % 2 == 1))

# === 昨日更新 ===
upd_start = rev_start + 2 + len(sorted_by_rev) + 1
ws2.merge_cells(f'A{upd_start}:F{upd_start}')
ws2.cell(row=upd_start, column=1, value='五、昨日服务更新记录').font = font_subtitle()

upd_headers = ['服务ID', '服务名称', '更新类型', '更新详情']
r = upd_start + 1
for j, h in enumerate(upd_headers):
    ws2.cell(row=r, column=j+1, value=h)
apply_header_row(ws2, r, 4)

for i, u in enumerate(svc_tracking['updates']):
    r = upd_start + 2 + i
    ws2.cell(row=r, column=1, value=u.get('service_id', ''))
    ws2.cell(row=r, column=2, value=u.get('name', ''))
    ws2.cell(row=r, column=3, value=u.get('type', ''))
    ws2.cell(row=r, column=4, value=u.get('detail', ''))
    apply_body_row(ws2, r, 4, alt=(i % 2 == 1))

# === 新增服务 ===
new_start = upd_start + 2 + len(svc_tracking['updates']) + 1
ws2.merge_cells(f'A{new_start}:F{new_start}')
ws2.cell(row=new_start, column=1, value='六、新增服务').font = font_subtitle()

new_headers = ['服务ID', '名称', '分类', '定价(ATEX)', '单位', '上架理由']
r = new_start + 1
for j, h in enumerate(new_headers):
    ws2.cell(row=r, column=j+1, value=h)
apply_header_row(ws2, r, 6)

for i, ns in enumerate(svc_tracking['new_services']):
    r = new_start + 2 + i
    ws2.cell(row=r, column=1, value=ns.get('service_id', ''))
    ws2.cell(row=r, column=2, value=ns.get('name', ''))
    ws2.cell(row=r, column=3, value=ns.get('category', ''))
    ws2.cell(row=r, column=4, value=ns.get('price', ''))
    ws2.cell(row=r, column=5, value=ns.get('unit', ''))
    ws2.cell(row=r, column=6, value=ns.get('rationale', ''))
    apply_body_row(ws2, r, 6, alt=(i % 2 == 1))

# === 下架记录 ===
paused = [u for u in svc_tracking['updates'] if u.get('type') == 'status_change']
del_start = new_start + 2 + len(svc_tracking['new_services']) + 1
ws2.merge_cells(f'A{del_start}:F{del_start}')
ws2.cell(row=del_start, column=1, value='七、下架/暂停记录').font = font_subtitle()

if paused:
    del_headers = ['服务ID', '服务名称', '变更类型', '详情']
    r = del_start + 1
    for j, h in enumerate(del_headers):
        ws2.cell(row=r, column=j+1, value=h)
    apply_header_row(ws2, r, 4)
    for i, p in enumerate(paused):
        r = del_start + 2 + i
        ws2.cell(row=r, column=1, value=p.get('service_id', ''))
        ws2.cell(row=r, column=2, value=p.get('name', ''))
        ws2.cell(row=r, column=3, value=p.get('type', ''))
        ws2.cell(row=r, column=4, value=p.get('detail', ''))
        apply_body_row(ws2, r, 4, alt=(i % 2 == 1))
else:
    ws2.cell(row=del_start+1, column=1, value='无下架/暂停记录').font = font_note()

# === 待办行动项 ===
act_start = del_start + 2 + len(paused) + 1
ws2.merge_cells(f'A{act_start}:F{act_start}')
ws2.cell(row=act_start, column=1, value='八、待办行动项').font = font_subtitle()

for i, item in enumerate(svc_tracking['action_items'][:15]):
    r = act_start + 1 + i
    ws2.cell(row=r, column=1, value=f'{i+1}')
    ws2.merge_cells(f'B{r}:F{r}')
    ws2.cell(row=r, column=2, value=item)
    ws2.cell(row=r, column=1).font = font_body()
    ws2.cell(row=r, column=2).font = font_body()
    ws2.cell(row=r, column=1).alignment = align_center
    ws2.cell(row=r, column=2).alignment = align_left
    # Highlight pricing suggestions
    if '降价' in item or '偏高' in item:
        ws2.cell(row=r, column=2).font = font_highlight()

# 列宽
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 50

# ============================================================
# Sheet3: 交易平台运营数据
# ============================================================
ws3 = wb.create_sheet("交易平台运营数据")

# 读取数据
with open('/home/z/my-project/token_exchange/data/daily_platform_ops.json') as f:
    platform_ops = json.load(f)
with open('/home/z/my-project/token_exchange/data/bootstrap_report.json') as f:
    bootstrap = json.load(f)

# 标题
ws3.merge_cells('A1:F1')
ws3['A1'] = 'ATEX交易平台运营数据 | 2026-05-27'
ws3['A1'].font = font_title()
ws3['A1'].alignment = align_center

ws3.merge_cells('A2:F2')
ws3['A2'] = f'数据采集时间：{platform_ops.get("generated_at", "N/A")} | 来源：{platform_ops.get("source", "N/A")}'
ws3['A2'].font = font_note()
ws3['A2'].alignment = align_center

# === Token交易数据 ===
ws3.merge_cells('A4:F4')
ws3['A4'] = '一、Token交易数据'
ws3['A4'].font = font_subtitle()

trading = platform_ops.get('trading', {})
token_data = [
    ['指标', '数值'],
    ['当日交易量(ATEX)', f'{trading.get("daily_volume", 0)}'],
    ['当日交易笔数', f'{trading.get("daily_trades", 0)}'],
    ['最新成交价', f'{trading.get("last_price", "N/A")}'],
    ['当日最高价', f'{trading.get("daily_high", "N/A")}'],
    ['当日最低价', f'{trading.get("daily_low", "N/A")}'],
    ['挂单买单数', f'{trading.get("open_bids", 0)}'],
    ['挂单卖单数', f'{trading.get("open_asks", 0)}'],
    ['历史总交易笔数', f'{trading.get("total_trades_all", 0)}'],
    ['历史总佣金(ATEX)', f'{trading.get("total_commission_earned", 0):.2f}'],
]

for i, row_data in enumerate(token_data):
    r = 5 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))

# === 订单簿深度 ===
ob_start = 5 + len(token_data) + 1
ws3.merge_cells(f'A{ob_start}:F{ob_start}')
ws3.cell(row=ob_start, column=1, value='二、订单簿深度').font = font_subtitle()

ob_headers = ['方向', '价格(ATEX)', '数量(ATEX)', '订单数']
r = ob_start + 1
for j, h in enumerate(ob_headers):
    ws3.cell(row=r, column=j+1, value=h)
apply_header_row(ws3, r, 4)

bids = trading.get('orderbook_bids', [])
asks = trading.get('orderbook_asks', [])
r = ob_start + 2
for bid in bids:
    ws3.cell(row=r, column=1, value='买单(Bid)')
    ws3.cell(row=r, column=2, value=bid.get('price', 0))
    ws3.cell(row=r, column=3, value=bid.get('amount', 0))
    ws3.cell(row=r, column=4, value=bid.get('orders', 1))
    apply_body_row(ws3, r, 4)
    for c in range(1, 5):
        ws3.cell(row=r, column=c).fill = fill_green()
    r += 1

for ask in asks:
    ws3.cell(row=r, column=1, value='卖单(Ask)')
    ws3.cell(row=r, column=2, value=ask.get('price', 0))
    ws3.cell(row=r, column=3, value=ask.get('amount', 0))
    ws3.cell(row=r, column=4, value=ask.get('orders', 1))
    apply_body_row(ws3, r, 4)
    for c in range(1, 5):
        ws3.cell(row=r, column=c).fill = fill_red()
    r += 1

# 买卖价差
if bids and asks:
    spread = asks[0]['price'] - bids[0]['price']
    ws3.cell(row=r, column=1, value='买卖价差')
    ws3.cell(row=r, column=2, value=f'{spread:.2f} ({spread/asks[0]["price"]*100:.1f}%)')
    ws3.cell(row=r, column=1).font = font_highlight()
    ws3.cell(row=r, column=2).font = font_highlight()

# === 服务交易数据 ===
svc_trade_start = r + 2
ws3.merge_cells(f'A{svc_trade_start}:F{svc_trade_start}')
ws3.cell(row=svc_trade_start, column=1, value='三、服务交易数据').font = font_subtitle()

sm = platform_ops.get('service_market', {})
svc_trade_data = [
    ['指标', '数值'],
    ['服务总数', f'{sm.get("total_services", 0)}'],
    ['当日服务购买笔数', f'{sm.get("total_service_orders_today", 0)}'],
    ['当日服务交易额(ATEX)', f'{sm.get("total_service_volume_today", 0):.1f}'],
    ['历史总订单数', f'{sm.get("all_time_orders", 0)}'],
    ['历史总交易额(ATEX)', f'{sm.get("all_time_volume", 0):.1f}'],
    ['历史总佣金(ATEX)', f'{sm.get("all_time_commission", 0):.1f}'],
    ['平均服务价格(ATEX)', f'{sm.get("avg_service_price", 0):.2f}'],
]

for i, row_data in enumerate(svc_trade_data):
    r = svc_trade_start + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))

# === 热门服务 ===
hot_start = svc_trade_start + 1 + len(svc_trade_data) + 1
ws3.merge_cells(f'A{hot_start}:F{hot_start}')
ws3.cell(row=hot_start, column=1, value='四、热门服务（收入排行Top 5）').font = font_subtitle()

hot_headers = ['排名', '服务名称', '分类', '销量', '收入(ATEX)']
r = hot_start + 1
for j, h in enumerate(hot_headers):
    ws3.cell(row=r, column=j+1, value=h)
apply_header_row(ws3, r, 5)

top_svcs = sm.get('top_services_by_revenue', [])
for i, s in enumerate(top_svcs):
    r = hot_start + 2 + i
    ws3.cell(row=r, column=1, value=i+1)
    ws3.cell(row=r, column=2, value=s.get('name', ''))
    ws3.cell(row=r, column=3, value=s.get('category', ''))
    ws3.cell(row=r, column=4, value=s.get('total_sold', 0))
    ws3.cell(row=r, column=5, value=s.get('revenue', 0))
    apply_body_row(ws3, r, 5, alt=(i % 2 == 1))

# === 用户数据 ===
user_start = hot_start + 2 + len(top_svcs) + 1
ws3.merge_cells(f'A{user_start}:F{user_start}')
ws3.cell(row=user_start, column=1, value='五、用户数据').font = font_subtitle()

users = platform_ops.get('users', {})
user_data = [
    ['指标', '数值'],
    ['总账户数', f'{users.get("total_accounts", 0)}'],
    ['当日新增用户', f'{users.get("new_accounts_today", 0)}'],
    ['虚拟账户数', f'{users.get("virtual_accounts", 0)}'],
    ['真实账户数', f'{users.get("total_accounts", 0) - users.get("virtual_accounts", 0)}'],
    ['虚拟/真实比', f'{users.get("virtual_accounts", 0)}:{users.get("total_accounts", 0) - users.get("virtual_accounts", 0)}'],
    ['虚拟占比', f'{users.get("virtual_accounts", 0)/max(users.get("total_accounts", 1), 1)*100:.1f}%'],
    ['Maker佣金率', f'{users.get("commission_rates", {}).get("maker", "N/A")}'],
    ['Taker佣金率', f'{users.get("commission_rates", {}).get("taker", "N/A")}'],
]

for i, row_data in enumerate(user_data):
    r = user_start + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))

# === Provider收入排行 ===
prov_start = user_start + 1 + len(user_data) + 1
ws3.merge_cells(f'A{prov_start}:F{prov_start}')
ws3.cell(row=prov_start, column=1, value='六、Provider收入排行').font = font_subtitle()

prov_headers = ['排名', 'Provider', '收入(ATEX)']
r = prov_start + 1
for j, h in enumerate(prov_headers):
    ws3.cell(row=r, column=j+1, value=h)
apply_header_row(ws3, r, 3)

providers_earned = bootstrap.get('purchases', {}).get('providers_earned', {})
# Also get from all-time data
all_providers = {}
for order in platform_ops.get('recent_service_orders', []):
    prov = order.get('provider', '')
    cost = order.get('cost', 0)
    if prov:
        all_providers[prov] = all_providers.get(prov, 0) + cost

# Merge with bootstrap
for prov, earned in providers_earned.items():
    all_providers[prov] = all_providers.get(prov, 0) + earned

sorted_providers = sorted(all_providers.items(), key=lambda x: x[1], reverse=True)
for i, (prov, earned) in enumerate(sorted_providers):
    r = prov_start + 2 + i
    ws3.cell(row=r, column=1, value=i+1)
    ws3.cell(row=r, column=2, value=prov)
    ws3.cell(row=r, column=3, value=earned)
    apply_body_row(ws3, r, 3, alt=(i % 2 == 1))

# === 佣金数据 ===
comm_start = prov_start + 2 + len(sorted_providers) + 1
ws3.merge_cells(f'A{comm_start}:F{comm_start}')
ws3.cell(row=comm_start, column=1, value='七、佣金数据').font = font_subtitle()

finance = platform_ops.get('finance', {})
comm_data = [
    ['指标', '数值'],
    ['Token交易佣金(ATEX)', f'{finance.get("commission_token_trading", 0):.2f}'],
    ['服务交易佣金(ATEX)', f'{finance.get("commission_service_trading", 0):.2f}'],
    ['当日总佣金(ATEX)', f'{finance.get("total_commission_today", 0):.2f}'],
    ['累计总佣金(ATEX)', f'{finance.get("cumulative_commission", 0):.2f}'],
    ['当日Taker佣金', f'{finance.get("today_service_commission_breakdown", {}).get("taker", 0):.2f}'],
    ['当日Maker佣金', f'{finance.get("today_service_commission_breakdown", {}).get("maker", 0):.2f}'],
]

for i, row_data in enumerate(comm_data):
    r = comm_start + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))

# === 推广效果 ===
promo_start = comm_start + 1 + len(comm_data) + 1
ws3.merge_cells(f'A{promo_start}:F{promo_start}')
ws3.cell(row=promo_start, column=1, value='八、推广效果').font = font_subtitle()

promo = platform_ops.get('promotion', {})
promo_data = [
    ['指标', '数值'],
    ['最近推广日期', f'{promo.get("last_promotion_date", "N/A")}'],
    ['推广记录总数', f'{promo.get("total_promotions_logged", 0)}'],
    ['GitHub Stars', f'{promo.get("github_stars", 0)}'],
    ['GitHub Forks', f'{promo.get("github_forks", 0)}'],
    ['GitHub Open Issues', f'{promo.get("github_open_issues", 0)}'],
]

for i, row_data in enumerate(promo_data):
    r = promo_start + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))

# 渠道健康
ch_start = promo_start + 1 + len(promo_data) + 1
ws3.cell(row=ch_start, column=1, value='渠道健康状态：').font = font_subtitle()
channels = promo.get('channel_health', {})
for i, (ch, status) in enumerate(channels.items()):
    r = ch_start + 1 + i
    ws3.cell(row=r, column=1, value=ch)
    ws3.merge_cells(f'B{r}:F{r}')
    ws3.cell(row=r, column=2, value=status)
    ws3.cell(row=r, column=1).font = font_body()
    ws3.cell(row=r, column=2).font = font_body()

# === 安全事件 ===
sec_start = ch_start + 1 + len(channels) + 1
ws3.merge_cells(f'A{sec_start}:F{sec_start}')
ws3.cell(row=sec_start, column=1, value='九、安全事件').font = font_subtitle()

security = platform_ops.get('security', {})
sec_data = [
    ['指标', '数值'],
    ['当日安全事件', f'{security.get("incidents_today", 0)}'],
    ['限频触发次数', f'{security.get("rate_limit_triggers", 0)}'],
    ['自交易拦截次数', f'{security.get("self_trade_blocks", 0)}'],
    ['价格偏离熔断次数', f'{security.get("price_deviation_circuit_breakers", 0)}'],
]

for i, row_data in enumerate(sec_data):
    r = sec_start + 1 + i
    for j, val in enumerate(row_data):
        ws3.cell(row=r, column=j+1, value=val)
    if i == 0:
        apply_header_row(ws3, r, 2)
    else:
        apply_body_row(ws3, r, 2, alt=(i % 2 == 0))
        if isinstance(val, (int, float)) and val > 0:
            ws3.cell(row=r, column=2).font = font_highlight()

# === 待办行动项 ===
act3_start = sec_start + 1 + len(sec_data) + 1
ws3.merge_cells(f'A{act3_start}:F{act3_start}')
ws3.cell(row=act3_start, column=1, value='十、待办行动项').font = font_subtitle()

platform_actions = [
    f'1. Token交易量连续多日为零，需推动做市商活动或调整挂单价格（当前Bid 1.0/Ask 1.5，价差50%）',
    f'2. 虚拟账户占比{users.get("virtual_accounts", 0)/max(users.get("total_accounts", 1), 1)*100:.0f}%，需加速真实用户获取',
    f'3. GitHub仅{promo.get("github_stars", 0)} Star/{promo.get("github_forks", 0)} Fork，推广效果待提升',
    f'4. 最近推广日期{promo.get("last_promotion_date", "N/A")}，已超过一周未执行推广',
    f'5. 服务交易当日{sm.get("total_service_orders_today", 0)}笔，均为虚拟买家，需引入真实Agent用户',
    f'6. 累计佣金{finance.get("cumulative_commission", 0):.1f} ATEX，距离盈利目标差距大',
    f'7. Provider邀请5类均pending状态，需主动触达AI模型/数据/安全/开发者/内容提供商',
    f'8. 安全事件为0，系统稳定但需持续监控MCP安全风险',
]

for i, action in enumerate(platform_actions):
    r = act3_start + 1 + i
    ws3.merge_cells(f'A{r}:F{r}')
    ws3.cell(row=r, column=1, value=action)
    ws3.cell(row=r, column=1).font = font_body()
    ws3.cell(row=r, column=1).alignment = align_left

# 列宽
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 50

# ===== 保存 =====
output_path = '/home/z/my-project/reports/ATEX综合日报_20260527.xlsx'
wb.save(output_path)
print(f'Report saved to {output_path}')
